"""
Generate Excel template with dropdown menus for ComStock Retrofit Planner
Uses openpyxl to create real Excel data validation dropdowns
"""

import json
import csv
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_template_with_dropdowns(output_filename='comstock_input_template_with_dropdowns.xlsx'):
    """Generate Excel template with dropdown menus for each column"""
    
    print("Loading column options...")
    with open('column_options.json', 'r') as f:
        column_options = json.load(f)
    
    print(f"Loaded {len(column_options)} columns with options from dataset")
    
    # Read the actual input_data.csv to get the exact columns we need
    print("Reading input_data.csv to get required columns...")
    with open('input_data.csv', 'r') as f:
        csv_reader = csv.reader(f)
        all_columns = next(csv_reader)  # Get header row
    
    print(f"Found {len(all_columns)} columns in input_data.csv")
    
    # Only the last 2 columns are outputs, all others are inputs
    input_columns = all_columns[:-2]  # First 48 columns
    output_columns = all_columns[-2:]  # Last 2 columns
    
    print(f"  - {len(input_columns)} input columns (user fillable)")
    print(f"  - {len(output_columns)} output columns (model predictions)")
    
    # Create workbook and main sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "ComStock_Input_Template"
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    print("Writing headers...")
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Add 20 empty rows for data entry
    for row_idx in range(2, 22):
        for col_idx in range(1, len(all_columns) + 1):
            ws.cell(row=row_idx, column=col_idx).value = ""
    
    # Add data validation (dropdowns) for input columns
    print("Adding dropdown validation to columns...")
    
    columns_with_dropdowns = 0
    columns_with_comments = 0
    
    for col_idx, col_name in enumerate(all_columns, start=1):
        # Only add dropdowns for input columns that have options in our dataset
        if col_name not in input_columns or col_name not in column_options:
            continue
            
        options = column_options[col_name]
        col_letter = get_column_letter(col_idx)
        
        # Excel data validation has a limit of 255 characters for the list formula
        # For columns with many options or long strings, we'll skip dropdowns
        options_string = ','.join([str(opt) for opt in options])
        
        if len(options) <= 100 and len(options_string) <= 255:
            # Create dropdown with direct list
            dv = DataValidation(
                type="list",
                formula1=f'"{options_string}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid Input",
                error=f"Please select a value from the dropdown or enter a custom value"
            )
            dv.prompt = f"Select from {len(options)} available options or enter custom value"
            dv.promptTitle = col_name
            
            # Apply to all data rows (rows 2-1000)
            dv.add(f"{col_letter}2:{col_letter}1000")
            ws.add_data_validation(dv)
            columns_with_dropdowns += 1
            print(f"  ✓ Added dropdown for {col_name} ({len(options)} options)")
            
        elif len(options) <= 1000:
            # For larger lists, create a hidden reference sheet
            # Create or get the Options sheet
            if "Options_Lists" not in wb.sheetnames:
                ws_options = wb.create_sheet("Options_Lists")
                ws_options.sheet_state = 'hidden'
            else:
                ws_options = wb["Options_Lists"]
            
            # Find the next available column in Options sheet
            options_col = ws_options.max_column + 1
            options_col_letter = get_column_letter(options_col)
            
            # Write the column name in row 1
            ws_options.cell(row=1, column=options_col).value = col_name
            
            # Write all options to the Options sheet
            for opt_idx, opt in enumerate(options, start=2):
                ws_options.cell(row=opt_idx, column=options_col).value = opt
            
            # Create data validation referring to the Options sheet
            dv = DataValidation(
                type="list",
                formula1=f"=Options_Lists!${options_col_letter}$2:${options_col_letter}${len(options)+1}",
                allow_blank=True,
                showErrorMessage=False
            )
            dv.prompt = f"Select from {len(options)} available options or enter custom value"
            dv.promptTitle = col_name
            
            # Apply to all data rows
            dv.add(f"{col_letter}2:{col_letter}1000")
            ws.add_data_validation(dv)
            columns_with_dropdowns += 1
            print(f"  ✓ Added reference dropdown for {col_name} ({len(options)} options)")
            
        else:
            # Too many options - add a note instead
            cell = ws.cell(row=1, column=col_idx)
            cell.comment = None  # Remove any existing comment
            from openpyxl.comments import Comment
            comment_text = f"{len(options)} options available.\nSample: {', '.join([str(opt) for opt in options[:5]])}...\n\nUse Column Options Explorer to see all options or enter your own value."
            cell.comment = Comment(comment_text, "System")
            columns_with_comments += 1
            print(f"  ⓘ Added comment for {col_name} ({len(options)} options - too many for dropdown)")
    
    # Mark output columns
    print("Marking output columns...")
    for col_name in output_columns:
        col_idx = all_columns.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        
        # Add gray background to output columns
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        for row_idx in range(2, 22):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = gray_fill
            if row_idx == 2:
                cell.value = "(Model Output - Leave Empty)"
    
    # Create an Instructions sheet at the end (so ComStock_Input_Template remains first)
    print("Creating instructions sheet...")
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80
    
    instructions = [
        ["✨ ComStock Retrofit Planner - Input Template with Dropdown Menus"],
        [""],
        ["📋 HOW TO USE THIS TEMPLATE:"],
        [""],
        ["1. Go to the 'ComStock_Input_Template' sheet"],
        ["2. Click on any cell in a column with a dropdown menu"],
        ["3. You'll see a dropdown arrow - click it to select from available options"],
        ["4. Or type your own custom value directly into the cell"],
        ["5. Fill in data for all input columns (leave output columns empty)"],
        ["6. You can add more rows as needed (up to 1000 buildings)"],
        ["7. Save the file and upload it to the Retrofit Planner"],
        [""],
        ["💡 TIPS:"],
        [""],
        ["• Columns with ≤1000 options have automatic dropdown menus"],
        ["• Columns with >1000 options have comments with sample values"],
        ["• You can always enter custom values - the model is flexible"],
        ["• Gray columns are outputs - the model will fill these automatically"],
        ["• Use the 'Column Options Explorer' button on the website to browse all options"],
        [""],
        [f"📊 TEMPLATE STATISTICS:"],
        [""],
        [f"• Total Columns: {len(all_columns)}"],
        [f"• Input Columns: {len(input_columns)}"],
        [f"• Output Columns: {len(output_columns)}"],
        [f"• Columns with Dropdowns: {columns_with_dropdowns}"],
        [f"• Columns with Comments: {columns_with_comments}"],
        [f"• Empty Rows Provided: 20 (add more as needed)"],
        [""],
        ["✅ Ready to start? Go to the ComStock_Input_Template sheet!"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=1):
        cell = ws_instructions.cell(row=row_idx, column=1)
        cell.value = row_data[0]
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="4472C4")
        elif row_data[0].startswith(('📋', '💡', '📊', '✅')):
            cell.font = Font(bold=True, size=12)
    
    # Save the workbook
    print(f"Saving workbook to {output_filename}...")
    wb.save(output_filename)
    print(f"✅ Template generated successfully: {output_filename}")
    print(f"   - {len(all_columns)} total columns ({len(input_columns)} input + {len(output_columns)} output)")
    print(f"   - {columns_with_dropdowns} columns with dropdown menus")
    print(f"   - {columns_with_comments} columns with cell comments")
    print(f"   - 20 empty rows ready for data entry")

if __name__ == "__main__":
    generate_template_with_dropdowns()
